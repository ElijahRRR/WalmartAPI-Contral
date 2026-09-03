"""绩效归因:问题描述必须说「为什么违规」,不许再是明细的拍平版。

所有者 2026-09-03 报的形态:订单中心绩效表里「问题描述」与「明细」一字不差
(PO/Item ID/Category/GMV loss/…),看不出这单到底是无扫描还是追踪异常。
"""

import json

from api import insights
from services import kpi, perf_reason

# 所有者贴的真实行(取消率报表,PO 129124487111861)——回归基准就用它
OWNER_ROW = {
    "PO #": "129124487111861", "Item ID": "20852409198", "Category": "PLUMBING",
    "GMV loss": "47.55", "Order date": "2026-08-31", "Order Line #": "1",
    "Sales Order #": "200014995969280", "Item Condition": "New",
    "Cancellation timestamp": "2026-09-01 01:34:54.000 UTC",
}


def test_owner_row_is_no_longer_the_detail_flattened():
    """问题描述 ≠ 明细:原因在前,只带两三条佐证,不复述整行。"""
    desc = perf_reason.describe("cancellations", "Out of Stock", OWNER_ROW)
    assert desc == "缺货取消 · 损失 $47.55 / 取消于 09-01 01:34 UTC"
    # 明细里那些跟"为什么"无关的列一个都不许再出现在问题描述里
    for noise in ("PO #", "Item ID", "PLUMBING", "Sales Order #", "New",
                  "Order Line"):
        assert noise not in desc
    assert desc != "; ".join(f"{k}:{v}" for k, v in OWNER_ROW.items())


def test_vtr_official_buckets_become_chinese_reasons():
    """VTR 五个官方 seller-accountable 桶(marketplacelearn 2026-09-03 核对)。"""
    row = {"PO #": "1", "Carrier": "FedEx", "Tracking Number": "3928392839283"}
    got = {b: perf_reason.classify("vtr", b, row)[0] for b in (
        "No Carrier Scan", "Invalid Tracking ID", "Invalid tracking URL",
        "Misleading Tracking", "Non-Integrated Carrier")}
    assert got == {"No Carrier Scan": "承运商无扫描",
                   "Invalid Tracking ID": "追踪号无效",
                   "Invalid tracking URL": "追踪链接无效",
                   "Misleading Tracking": "追踪信息不实",
                   "Non-Integrated Carrier": "承运商未对接"}
    # 佐证带上承运商与单号:VTR 的排查抓手就是这两个
    assert perf_reason.describe("vtr", "No Carrier Scan", row) == \
        "承运商无扫描 · 承运商 FedEx / 单号 3928392839283"


def test_row_reason_fills_in_when_the_sheet_name_says_nothing():
    """sheet 只写「Seller Accountable」时不含原因,靠行内原因列。"""
    row = {"PO #": "1", "Defect reason": "Invalid tracking ID"}
    assert perf_reason.classify("vtr", "Seller Accountable", row)[0] == "追踪号无效"


def test_sheet_bucket_and_row_reason_are_both_kept_when_they_differ():
    """sheet 桶 = 记在哪一类缺陷上,行内原因 = 具体发生了什么;实测会不同。"""
    assert perf_reason.classify("returns", "Incorrect item", {
        "Return reason description": "Not as described/pictured"})[0] == \
        "错发商品(与描述/图片不符)"
    # 一样就不重复说
    assert perf_reason.classify("returns", "Damaged", {
        "Return reason description": "Damaged"})[0] == "商品破损"
    # 补语等于没说时也不挂上去
    assert perf_reason.classify("refunds", "Incorrect item", {
        "Refund reason": "Miscellaneous"})[0] == "错发商品"


def test_product_category_column_is_not_a_reason():
    """Category:PLUMBING 是商品类目,不是违规原因——收错列会把类目写成原因。"""
    assert "PLUMBING" not in (perf_reason.describe("cancellations", "", OWNER_ROW) or "")


def test_unknown_bucket_passes_through_and_is_counted():
    """词表对不上时原样透传并计数,等人校准——不许静默归到「其他」。"""
    seen: dict[str, int] = {}
    desc = perf_reason.describe("cancellations", "Brand New Bucket", OWNER_ROW,
                                unknown_seen=seen)
    assert desc.startswith("Brand New Bucket · ")
    assert seen == {"Brand New Bucket": 1}


def test_order_api_fills_the_reason_the_report_withholds():
    """③ 订单库补全:report 不给原因,orders 的 cancellationReason 顶上并注明出处。"""
    desc = perf_reason.describe("cancellations", "Seller Accountable", OWNER_ROW,
                                ctx={"cancel_reason": "Out of stock"})
    assert desc.startswith("缺货取消(订单接口) · ")
    # 补全值也对不上词表时,原文照进,仍然注明出处
    assert perf_reason.classify(
        "cancellations", "", OWNER_ROW,
        ctx={"cancel_reason": "SOME_NEW_CODE"})[0] == "SOME_NEW_CODE(订单接口)"


def test_vtr_without_any_tracking_is_named_precisely():
    """报表和订单库都没有单号 = 根本没交追踪,比「追踪未通过校验」更准。"""
    assert perf_reason.classify("vtr", "", {"PO #": "1"})[0] == "未提供追踪号"
    # 库里有单号就不能这么说了,退回通用语
    assert perf_reason.classify("vtr", "", {"PO #": "1"},
                                ctx={"tracking_no": "TRK9"})[0] != "未提供追踪号"


def test_srr_inquiry_types_do_not_leak_into_cancellations():
    """SRR 的「Cancellation」是咨询类型,取消率报表里同名的不是。"""
    assert perf_reason.classify("srr", "Cancellation", {})[0] == "咨询:取消"
    reason, unknown = perf_reason.classify("cancellations", "Cancellation", {})
    assert reason == "Cancellation" and unknown == "Cancellation"


def test_not_accountable_rows_say_who_walmart_blamed():
    """不计入绩效的行也要说人话,不是空白。"""
    assert perf_reason.classify("otd", "", {}, accountable=False)[0] == \
        "送达超时但沃尔玛判非卖家责任"
    # 桶名本身就说明了原因时,以桶名为准(别再叠一句责任归属)
    assert perf_reason.classify("cancellations", "Customer requested", {},
                                accountable=False)[0] == "买家主动取消"


def test_every_metric_has_a_generic_sentence():
    """8 个指标都要有兜底话术——官方加指标时这条会先红。"""
    assert set(insights.METRICS) <= set(perf_reason._GENERIC)


def test_unknown_metric_returns_none_for_the_caller_to_fall_back():
    assert perf_reason.describe("自创指标", "", {"PO #": "1"}) is None


def test_evidence_is_capped_and_formatted():
    """佐证最多 3 条:问题描述是给人扫一眼的,不是第二份明细。列名用真实的。"""
    row = {"Delivered late by (Days)": "3", "Carrier": "USPS",
           "Shipping speed": "STANDARD", "Tracking number": "TRK1",
           "Expected delivery date": "2026-08-28",
           "Actual delivery timestamp": "2026-08-31 12:00:00 UTC"}
    desc = perf_reason.describe("otd", "Late Delivery", row)
    assert desc == "送达晚 · 迟 3 天 / 承运商 USPS / 单号 TRK1"
    assert desc.count(" / ") == 2


def test_a_date_column_never_lands_in_a_money_slot():
    """退款报表没有金额列,只有 GMV loss —— 宽 needle `refund` 会撞上
    `Refund Date`,实测写出过「退款 $2026-09-01」。"""
    desc = perf_reason.describe("refunds", "Lost", {
        "GMV loss": "38.47", "Category": "DISPOSABLE NAPKINS",
        "Refund Date": "2026-09-01"})
    assert desc == "包裹丢失 · 损失 $38.47 / 退款于 09-01"
    assert "$2026" not in desc


def test_actual_delivery_is_not_the_expected_one():
    """`delivery date` 是 `Expected delivery date` 的子串:实测「实际送达」
    抓成了「承诺送达」,两列打出同一个值。现在 OTD 佐证只用迟到天数+物流。"""
    desc = perf_reason.describe("otd", "Carrier Delays", {
        "Expected delivery date": "2026-09-01",
        "Actual delivery timestamp": "2026-09-03 04:12:00 UTC",
        "Delivered late by (Days)": "2", "Carrier": "China Post"})
    assert desc == "承运商延误 · 迟 2 天 / 承运商 China Post"


def test_parse_problem_report_writes_the_reason_not_the_flattened_row():
    """端到端:xlsx → 行,description 是归因,raw 仍是整行原文(一列不少)。"""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("No Carrier Scan")
    ws.append(["Sales Order #", "PO #", "Carrier", "Tracking Number"])
    ws.append(["SO1", "PO1", "FedEx", "TRK9"])
    buf = io.BytesIO()
    wb.save(buf)

    rows = kpi.parse_problem_report("vtr", buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["description"] == "承运商无扫描 · 承运商 FedEx / 单号 TRK9"
    assert json.loads(rows[0]["raw"])["Sales Order #"] == "SO1"   # 原文不丢


def test_parse_problem_report_uses_the_ctx_the_workflow_prepared():
    """perf_problems 备好的 {PO: 订单库补全} 要真的用上(报表不给原因那一档)。"""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Seller Accountable")
    ws.append(["Sales Order #", "PO #", "GMV loss"])
    ws.append(["SO1", "PO1", "47.55"])
    buf = io.BytesIO()
    wb.save(buf)

    rows = kpi.parse_problem_report("cancellations", buf.getvalue(),
                                    {"PO1": {"cancel_reason": "Out of stock"}})
    assert rows[0]["description"] == "缺货取消(订单接口) · 损失 $47.55"


def test_a_date_column_is_never_mistaken_for_a_reason():
    """"Issue date"/"Exception timestamp" 是时间列,不是原因列。"""
    row = dict(OWNER_ROW, **{"Issue date": "2026-09-01"})
    assert perf_reason.classify("cancellations", "Out of Stock", row)[0] == "缺货取消"


def test_evidence_does_not_repeat_the_reason():
    """原因原样透传时,值相同的佐证列不再复述同一句。"""
    desc = perf_reason.describe("negativeFeedback", "Negative Feedback", {
        "Feedback reason": "Broken on arrival", "Rating": "1",
        "Comments": "Broken on arrival"})
    assert desc == "Broken on arrival · 评分 1"


def test_row_reason_column_beats_a_reason_shaped_neighbour():
    """一张报表可能有多列名字里带 reason:必须命中指标专属的那一列。
    实测 INR 的 Return reason description 被泛化匹配抢走过。"""
    assert perf_reason.classify("itemNotReceived", "", {
        "Quantity not received": "1",
        "Return reason description": "Seller Issued Refund",
        "Some other reason": "Miscellaneous"})[0] == "卖家已主动退款"


# ── 生产语料回归(2026-09-03 M001 全 8 张报表 + 库内存量行实测)───────────────
# 这一组是真跑一轮拿到的**全部**桶名/原因取值,一条都不许回落成英文原文或
# 「报表未给原因」。官方改版式先在这里红。
PRODUCTION = [
    # (指标, sheet 名, 原因列取值, 期望中文原因)
    ("otd", "Carrier EDD later than promised", "", "承运商预计送达晚于承诺"),
    ("otd", "Carrier Delays", "", "承运商延误"),
    ("otd", "Late Shipment", "", "发货晚"),
    ("otd", "", "Carrier exception", "承运商异常"),
    ("cancellations", "Address is not serviceable", "", "误标地址不可送达"),
    ("cancellations", "Out-of-stock", "", "缺货取消"),
    ("cancellations", "Pricing errors", "", "定价错误取消"),
    ("cancellations", "Ship window expired", "", "超发货窗口取消"),
    ("cancellations", "", "Customer Cancellations", "买家取消"),
    ("vtr", "Misleading tracking", "", "追踪信息不实"),
    ("refunds", "Lost", "", "包裹丢失"),
    ("refunds", "Incorrect item", "", "错发商品"),
    ("refunds", "Defective", "", "商品有瑕疵"),
    ("refunds", "Damaged", "", "商品破损"),
    ("refunds", "", "Miscellaneous", "其他原因"),
    ("refunds", "", "Change_Mind Lower Price", "买家改主意:别处更便宜"),
    ("refunds", "", "Change_Mind No Longer Wanted", "买家改主意:不想要了"),
    ("returns", "Incorrect item", "", "错发商品"),
    ("returns", "Defective", "", "商品有瑕疵"),
    ("returns", "Damaged", "", "商品破损"),
    ("returns", "Arrived late", "", "到货晚"),
    ("returns", "", "Item arrived damaged", "到货破损"),
    ("returns", "", "Not as described/pictured", "与描述/图片不符"),
    ("returns", "", "No Longer Wanted", "买家不想要了"),
    ("returns", "", "Bought Somewhere Else", "已在别处买到"),
    ("itemNotReceived", "Lost", "", "包裹丢失"),
    ("itemNotReceived", "Item Missing", "", "商品缺失"),
    ("itemNotReceived", "", "Lost After Delivery", "妥投后丢失"),
    ("itemNotReceived", "", "Seller Issued Refund", "卖家已主动退款"),
    ("negativeFeedback", "Negative Feedback", "Item size or comfort", "尺寸或舒适度不合"),
    ("negativeFeedback", "Negative Feedback", "Item not as described", "与描述不符"),
    ("negativeFeedback", "Negative Feedback", "Missing item", "商品缺失"),
    ("srr", "Miscellaneous", "", "其他原因"),
    ("srr", "Track order", "", "咨询:查订单"),
]
_REASON_HEADER = {"otd": "Late Delivery Reason", "cancellations": "Cancellation reason",
                  "refunds": "Refund reason", "returns": "Return reason description",
                  "itemNotReceived": "Return reason description",
                  "negativeFeedback": "Feedback reason"}


def test_every_production_bucket_is_translated():
    seen: dict[str, int] = {}
    bad = []
    for metric, sheet, reason_val, want in PRODUCTION:
        row = {_REASON_HEADER[metric]: reason_val} if reason_val else {}
        got, miss = perf_reason.classify(metric, sheet, row)
        if got != want or miss:
            bad.append(f"{metric}/{sheet or '(空)'}/{reason_val or '-'}: {got!r}")
        perf_reason.describe(metric, sheet, row, unknown_seen=seen)
    assert not bad, "生产桶名归类不符:\n" + "\n".join(bad)
    assert seen == {}, f"生产语料里还有未收录桶名:{seen}"


def test_na_and_negative_feedback_sheet_carry_no_reason():
    """「N/A」「Negative Feedback」不是原因:必须继续降级,别写进问题描述。"""
    assert perf_reason.classify("returns", "", {
        "Return reason description": "N/A"}, accountable=False)[0] == \
        "买家退货,沃尔玛判非卖家责任"
    assert perf_reason.classify("negativeFeedback", "Negative Feedback",
                                {})[0] == "买家差评"
