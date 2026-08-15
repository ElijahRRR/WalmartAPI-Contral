"""order_history_import — 旧订单汇总 xlsx → orders.order_lines(历史补全,可重跑)。

用法:
  python cli.py order_history_import -p file=<xlsx路径>              # 预览
  python cli.py order_history_import -p file=<xlsx路径> -p apply=1   # 真正入库
  python cli.py order_history_import -p file=<...> -p sheet=合并总表  # 换工作表
  python cli.py order_history_import -p file=<...> -p since=2026-02-15  # 只导某日之后

数据源:所有者提供的「订单数据合并」workbook,`合并总表` 一行一个订单行,
15 万行跨 2024-03 ~ 2026-07。**只取销售侧六列**(所有者定稿 2026-08-15):
统一订单日期 / 店铺名称 / SKU / 商品标题 / 数量 / 销售额USD;
采购成本、利润估算、退款原因、统计状态四列**不导**。

身份与防重:
- `po_id` = 合并键**去掉末尾的 SKU 再 strip**。⚠ 不是"前 15 位":生产文件里
  有 PO 与 SKU 之间夹 8 个空格的行(150,863 行中 1 行),按位切会切出带空格的
  PO。全量校验过「合并键必以自身 SKU 结尾」,0 例外,所以按后缀切是安全的。
- `order_line_id` 用 services.order_lines.make_order_line_id(po, sku) 推导 ——
  与 order_sync 同一算法,将来 API 拉到同一张单算出的 id 完全一致,
  **不会产生第二条**,而是安静地落到同一行上。
- `INSERT ... ON CONFLICT DO NOTHING`(不指定冲突目标,order_line_id 主键与
  UNIQUE(po_id, sku) 两条约束都覆盖):**库里已有的行一个字段都不碰**。

来源标记:写 `source='历史数据'`。这些是残缺行(无物流地址、无 raw、
销售状态一律 Delivered),`order_center_push` 据此不把它们推去飞书。
⚠ order_sync 覆盖同一行时会把 source 写回 NULL —— 那正是想要的:
API 拉到真行之后它就不再是历史行,自动回到推送流。

时区:表里是日期(00:00:00)没有时区。**按 UTC 零点入库**,不是中国时间零点——
CN 是 UTC+8,UTC 零点加 8 小时仍在同一日历日,于是这个日期在 UTC 和 CN 两种
时区下读出来都是同一天;换成 CN 零点则会在 UTC 下退成前一天。
"""

import logging
from datetime import datetime, timezone

from registry import db
from services import order_lines

DANGEROUS = False

logger = logging.getLogger("workflows.order_history_import")

_SHEET = "合并总表"
_BATCH = 5000

# 表头关键词 → 字段,**顺序即优先级**(与 services.kpi._HIST_HEADER_MAP 同款)。
# 按关键词不按列位:列序是人在 Excel 里随手能改的,表头才是契约。
_HEADER_MAP: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("merge_key", ("合并键",)),
    ("order_date", ("统一订单日期", "订单日期")),
    ("store", ("店铺名称", "店铺")),
    ("sku", ("sku",)),                      # 在 product_name 之前:"商品SKU"归 sku
    ("product_name", ("商品标题", "商品名", "标题")),
    ("qty", ("数量",)),
    ("product_amount", ("销售额",)),
)
_REQUIRED = tuple(f for f, _ in _HEADER_MAP)

_INSERT = """
INSERT INTO orders.order_lines
    (order_line_id, store, po_id, line_number, sku, product_name, qty,
     sale_status, order_date, product_amount, source)
VALUES (%(order_line_id)s, %(store)s, %(po_id)s, '', %(sku)s, %(product_name)s,
        %(qty)s, 'Delivered', %(order_date)s, %(product_amount)s, %(source)s)
ON CONFLICT DO NOTHING
"""

_EXISTING_SQL = ("SELECT order_line_id FROM orders.order_lines"
                 " WHERE order_line_id = ANY(%(ids)s::text[])")


def map_header(header: list) -> tuple[dict[int, str], list[str]]:
    """输入:表头行 → 输出:({列下标: 字段名}, 未映射表头)。关键词包含匹配(casefold)。"""
    mapping: dict[int, str] = {}
    taken: set[str] = set()
    unmapped: list[str] = []
    for i, h in enumerate(header):
        text = str(h or "").strip().casefold()
        if not text:
            continue
        for field, kws in _HEADER_MAP:
            if field not in taken and any(k in text for k in kws):
                mapping[i] = field
                taken.add(field)
                break
        else:
            unmapped.append(str(h).strip())
    return mapping, unmapped


def split_po(merge_key, sku) -> str | None:
    """输入:合并键 + SKU → 输出:采购订单号;合并键不以 SKU 结尾则 None。

    ⚠ 按**后缀**切不按位切:生产文件里有 PO 与 SKU 之间夹空格的行,
    切完再 strip。返回 None 的行由调用方计数并跳过,不许猜。
    """
    key, sku = str(merge_key or "").strip(), str(sku or "").strip()
    if not key or not sku or not key.endswith(sku):
        return None
    po = key[:-len(sku)].strip()
    return po or None


def _as_utc_midnight(v) -> datetime | None:
    """输入:单元格日期值 → 输出:带 UTC 时区的当日零点;解析不了返回 None。"""
    if isinstance(v, datetime):
        return v.replace(tzinfo=timezone.utc) if v.tzinfo is None else v
    s = str(v or "").strip()[:10].replace("/", "-")
    try:
        d = datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        return None
    return d.replace(tzinfo=timezone.utc)


def _num(v):
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def parse_rows(header: list, rows, since: str = "") -> tuple[list[dict], dict[str, int]]:
    """输入:表头 + 数据行迭代器 + 可选起始日期 → 输出:(待插入行, 各口径计数)。

    跳的每一类都单独计数:静默丢行会让"导入了 12 万条"看起来正常,
    而少的那 3 万条永远没人发现。
    """
    mapping, _ = map_header(header)
    out: list[dict] = []
    seen: set[str] = set()
    stat = dict.fromkeys(("total", "bad_key", "bad_date", "bad_num",
                          "before_since", "dup_in_file"), 0)
    for raw in rows:
        if not any(c is not None and str(c).strip() for c in raw):
            continue
        stat["total"] += 1
        rec = {f: (raw[i] if i < len(raw) else None) for i, f in mapping.items()}
        po = split_po(rec.get("merge_key"), rec.get("sku"))
        sku = str(rec.get("sku") or "").strip()
        if not po:
            stat["bad_key"] += 1
            continue
        od = _as_utc_midnight(rec.get("order_date"))
        if od is None:
            stat["bad_date"] += 1
            continue
        if since and od.strftime("%Y-%m-%d") < since:
            stat["before_since"] += 1
            continue
        qty, amount = _num(rec.get("qty")), _num(rec.get("product_amount"))
        if qty is None or amount is None:
            stat["bad_num"] += 1
            continue
        line_id = order_lines.make_order_line_id(po, sku)
        if line_id in seen:
            stat["dup_in_file"] += 1
            continue
        seen.add(line_id)
        out.append({"order_line_id": line_id, "store": str(rec.get("store") or "").strip(),
                    "po_id": po, "sku": sku,
                    "product_name": str(rec.get("product_name") or "").strip()[:500],
                    "qty": int(qty), "order_date": od,
                    "product_amount": amount,
                    "source": order_lines.HISTORY_SOURCE})
    return out, stat


def _read_sheet(path: str, sheet: str):
    """输入:xlsx 路径 + 工作表名 → 输出:(表头, 数据行迭代器)。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise LookupError(f"工作表「{sheet}」不存在,本文件有:{wb.sheetnames}")
    ws = wb[sheet]
    # 生产实证(2026-08-06 kpi 那次):有些 xlsx 把 dimension 声明成单格,
    # read_only 模式按声明只读出 1 行 → 必须 reset_dimensions 重扫真实行
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()
    it = ws.iter_rows(values_only=True)
    return list(next(it)), it


def _existing(conn, ids: list[str]) -> set[str]:
    """输入:连接 + order_line_id 列表 → 输出:库里已存在的那些。

    先查再插,不靠 rowcount —— executemany 的 rowcount 在各驱动下语义不一,
    而"新增几条/撞上几条"正是人眼闸门要看的数,报错了比不报更坏。
    """
    found: set[str] = set()
    with conn.cursor() as cur:
        for i in range(0, len(ids), 50000):
            cur.execute(_EXISTING_SQL, {"ids": ids[i:i + 50000]})
            found.update(r[0] for r in cur.fetchall())
    return found


def _summary(stat: dict, rows: list[dict], hit: set[str], unmapped: list[str]) -> str:
    new = [r for r in rows if r["order_line_id"] not in hit]
    dates = [r["order_date"] for r in rows]
    lines = [f"解析 {stat['total']} 行 → 可导 {len(rows)} 行"
             f"(库中已有 {len(hit)},待新增 {len(new)})"]
    skipped = ",".join(f"{k} {v}" for k, v in stat.items()
                       if k != "total" and v)
    lines.append(f"跳过:{skipped or '无'}")
    if dates:
        lines.append(f"日期跨度 {min(dates):%Y-%m-%d} ~ {max(dates):%Y-%m-%d},"
                     f"店铺 {len({r['store'] for r in rows})} 家")
    if unmapped:
        lines.append(f"⚠ 未映射表头(不影响导入,确认没漏列):{unmapped}")
    return "\n".join(lines)


def run(params: dict) -> str:
    """输入:params(file/sheet/apply/since)→ 输出:导入摘要。"""
    path = str(params.get("file", "")).strip()
    if not path:
        return "缺 -p file=<xlsx路径>"
    sheet = str(params.get("sheet", _SHEET))
    since = str(params.get("since", "")).strip()
    apply_ = str(params.get("apply", "")) in ("1", "true", "yes")

    header, it = _read_sheet(path, sheet)
    _, unmapped = map_header(header)
    missing = [f for f in _REQUIRED if f not in dict(map_header(header)[0]).values()]
    if missing:
        return (f"表头缺字段 {missing} —— 实际表头 {header};"
                f" 关键词映射见 _HEADER_MAP,对不上先校准映射再导")

    rows, stat = parse_rows(header, it, since)
    if not rows:
        return "解析结果为空,未写入\n" + _summary(stat, rows, set(), unmapped)

    with db.pg_conn() as conn:
        hit = _existing(conn, [r["order_line_id"] for r in rows])
        new = [r for r in rows if r["order_line_id"] not in hit]
        head = _summary(stat, rows, hit, unmapped)
        if not apply_:
            sample = [(r["po_id"], r["sku"], r["store"],
                       r["order_date"].strftime("%Y-%m-%d")) for r in new[:5]]
            return (f"DRY-RUN(-p apply=1 真写)\n{head}\n"
                    f"新增样本={sample}")
        done = 0
        with conn.cursor() as cur:
            for i in range(0, len(new), _BATCH):
                chunk = new[i:i + _BATCH]
                cur.executemany(_INSERT, chunk)
                done += len(chunk)
                logger.info("历史订单导入:已提交 %d/%d", done, len(new))
    return f"{head}\n已写入 {done} 行(source={order_lines.HISTORY_SOURCE},冲突行原样保留)"
